from attr import define
from django.shortcuts import render
from numpy import True_
from .models import Task
from .forms import TaskForm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import xlwt
from django.shortcuts import redirect
import xlsxwriter
import asposecells
#import pass1

from datetime import date
 
# Returns the current local date
today = date.today()


def login(request):

    global login1
    global counter1

    
    if request.method=='POST':

        counter1 = True

        login1 = request.POST.get('login')

        if login1.find("k") != -1:
            login1 = login1.replace(login1[1], "к")
        
        password = request.POST.get('password')

        response = redirect('/home')

        log_spis = ["2к9391", "2к9392", "2к9393", "2к9394", "2к9311", "1к9391", "1к9392", "1к9393", "1к9394", "1к9311", "0к9391", "0к9392", "0к9393", "0к9394", "0к9311", "9к9391", "9к9392", "9к9393", "9к9394"]

        pas_spis = ["Qwcverzx21", "Qwcverzx22", "Qwcverzx23", "Qwcverzx24", "Qwcverzx211", "Qwcverzx11", "Qwcverzx12", "Qwcverzx13", "Qwcverzx14", "Qwcverzx111", "Qwcverzx01", "Qwcverzx02", "Qwcverzx03", "Qwcverzx04", "Qwcverzx011", "Qwcverzx91", "Qwcverzx92", "Qwcverzx93", "Qwcverzx94"]
  

        #log_spis = pass1.log_spis

        #pas_spis = pass1.pas_spis
        
        if (login1 in log_spis) and (password in pas_spis):
            
            i = log_spis.index(login1)
            j = pas_spis.index(password)

            if i == j:
                
                return response
                        
            if i != j:
                return render(request, 'main/login.html')

        else:
            return render(request, 'main/login.html')

        
        
    return render(request, 'main/login.html')
    

    
        



def index(request):


    

    file_name = "C:/Work/Python/Django/taskmanager/main/Missing.xlsx"
    wb = load_workbook(file_name)
    ws = wb.worksheets[0]

    #try: login1
    #except NameError: return redirect('/login')

    
        

    
        
    
    ms = ['spravka', 'zayavlenie', 'drugoe', 'neuvaj']
    if request.method=='POST':


        

        
            

        

        fio = request.POST.get('fio')
        
        zayavlenie = request.POST.get('zayava')
        
        drugoe = request.POST.get('drug')
    
        miss = request.POST.getlist('miss')

        #if isinstance(counter1, int):
        #    return response_log


        try: login1
        except NameError: return redirect('/login')

        group = login1

        if fio == '':
            return render(request, 'main/index.html')

        


            
        if group.find("k") != -1:
            group = group.replace(group[1], "к")

        
        st = "A2"
        i = 3
        while (ws[st].value) != group:
            st = "A" + str(i)
            i = i+1

        

        if miss == ["spravka"]:
           
            st =  st.replace(st[0], "B")
            if ws[st].value != None:
                ws[st] = "\n" + str(ws[st].value) + fio + "; " 
            if ws[st].value == None:
                ws[st] = fio + "; " 

        if miss == ["neuvaj"]:
           
            st =  st.replace(st[0], "E")
            if ws[st].value != None:
                ws[st] = "\n" + str(ws[st].value) + fio + "; " 
            if ws[st].value == None:
                ws[st] = fio + "; "
            
            #wb.save(file_name) 

        if miss == ["zayavlenie"]:

            if zayavlenie == '':
                return render(request, 'main/index.html')
            
            st =  st.replace(st[0], "C")
            if ws[st].value != None:
                ws[st] = "\n" + str(ws[st].value) + fio + " (" + str(zayavlenie) + "); "
            if ws[st].value == None:
                ws[st] =  fio + " (" + str(zayavlenie) + "); "
            
            #wb.save(file_name)  

        if miss == ["drugoe"]:

            if miss == '':
                return render(request, 'main/index.html')

            st =  st.replace(st[0], "D")
            if ws[st].value != None:
                ws[st] = "\n" + str(ws[st].value) + fio + " (" + str(drugoe) + "); " 
            if ws[st].value == None:
                ws[st] =  fio + " (" + str(drugoe) + "); " 
            

            
        wb.save(file_name)   

        #wb = Workbook("C:/Work/Python/Django/taskmanager/main/" + str(today) +".xlsx")
            

    return render(request, 'main/index.html')