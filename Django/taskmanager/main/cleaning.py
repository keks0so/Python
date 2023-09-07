from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import shutil
import schedule
import smtplib
import os
import time
import mimetypes
from pyfiglet import Figlet
from tqdm import tqdm
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.audio import MIMEAudio
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
import smtplib
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email import encoders


def send_email(text=None, template=None):
    today = date.today()


    file_name='C:/Work/Python/Django/taskmanager/main/Copy' + str(today)+ '.xlsx'




    message = MIMEMultipart()
    message["from"] = "denpirogoff@gmail.com"
    message["to"] = "okt@mrk-bsuir.by"
    message["subject"] = "Таблица по отсутствующим"



    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(file_name, "rb").read())
    encoders.encode_base64(part)
        
    part.add_header('Content-Disposition', 'attachment; filename= Copy' + str(today)+ '.xlsx')


    message.attach(MIMEText("Табличка", ))
    message.attach(part)


    with smtplib.SMTP(host="smtp.gmail.com", port=587) as smtp:

        smtp.ehlo()
        smtp.starttls()
        smtp.login("denpirogoff@gmail.com", "spinsttodabsqqic")
        smtp.send_message(message)



def res_cleaning():

    today = date.today()
    
    file_name = "C:/Work/Python/Django/taskmanager/main/Missing.xlsx"
    wb = load_workbook(file_name)
    ws = wb.worksheets[0]

   

    #df1 = pd.DataFrame([['122', str(122 - missed_people_spravka - missed_people_zayavlenie - missed_people_drugoe), str(missed_people_spravka), str(missed_people_zayavlenie), str(missed_people_drugoe)]],
    #                columns=['всего', 'отсутсв', 'справка', 'заявл', 'др'])
    #df1.to_excel("C:/Work/Python/Django/taskmanager/main/Res" + str(today) +".xlsx")


    original = r'C:/Work/Python/Django/taskmanager/main/Missing.xlsx'
    target = r'C:/Work/Python/Django/taskmanager/main/Copy' + str(today)+ '.xlsx'

    shutil.copyfile(original, target)






    #file_name = "C:/Work/Python/Django/taskmanager/main/Missing.xlsx"
    #wb = load_workbook(file_name)
    #ws = wb.worksheets[0]

    st = "A2"
    st2 = "B2"
    st3 = "C2"
    st4 = "D2"
    st5 = "E2"

    i = 3
    
    while ws[st].value != None:

        st = "A" + str(i)
        i = i + 1
        
    i = i-1
    

    
    st2 =  "B" + str(i)
    st3 =  "C" + str(i)
    st4 =  "D" + str(i)
    st5 =  "E" + str(i)

    i = i-1
    
    

    while st2 != "B1":
        if ws[st2].value != None:
            ws[st2] = ""
            
        
        if ws[st3].value != None:
            ws[st3] = ""
            
        
        if ws[st4].value != None:
            ws[st4] = ""


        if ws[st5].value != None:
            ws[st5] = ""
            

        
        st2 =  "B" + str(i)
        st3 =  "C" + str(i)
        st4 =  "D" + str(i)
        st5 =  "E" + str(i)

        i = i - 1

        

        wb.save(file_name)

def main():

    schedule.every().day.at('23:59').do(res_cleaning)

    schedule.every().day.at('08:30').do(send_email)
    while True:
        schedule.run_pending()

if __name__ == "__main__":
    main()





