import email
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import xlwt
import xlsxwriter
import asposecells
import win32com.client
import time
import os
import shutil
import subprocess
import ssl
import smtplib
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
from email.mime.base import MIMEBase
from email import encoders


login1 = input("Log: ")
password = input("Pass: ")


if login1.find("k") != -1:
        login1 = login1.replace(login1[1], "к")
        
   

log_spis = ["2к9391", "2к9392", "2к9393", "2к9394", "2к9311", "1к9391", "1к9392", "1к9393", "1к9394", "1к9311", "0к9391", "0к9392", "0к9393", "0к9394", "0к9311", "9к9391", "9к9392", "9к9393", "9к9394"]

pas_spis = ["2к9391", "2к9392", "2к9393", "2к9394", "2к9311", "1к9391", "1к9392", "1к9393", "1к9394", "1к9311", "0к9391", "0к9392", "0к9393", "0к9394", "0к9311", "9к9391", "9к9392", "9к9393", "9к9394"]
        
if (login1 in log_spis) and (password in pas_spis):
            
    i = log_spis.index(login1)
    j = pas_spis.index(password)

    if i == j:
        print("good")
                
    if i != j:
        print("Не правельный логин или пароль")

else:
    print("bad")
                








