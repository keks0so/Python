from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import shutil
import schedule
import smtplib
import os
import time
import mimetypes
from pyfiglet import Figlet
from tqdm import tqdm
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.audio import MIMEAudio
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from platform import python_version


def send_email():
    
    today = date.today()

    server = 'smtp.mail.ru'
    #server = 'smtp.gmail.com'
    user = 'denpirogoff@gmail.com'
    password = 'gyt67ujnB3'

    recipients = 'denpiro444@gmail.com'
    sender = 'denpirogoff@gmail.com'
    subject = 'Таблица отсутствующих за ' + str(today)
    

    filepath = 'C:/Work/Python/Django/taskmanager/main/Copy' + str(today)+ '.xlsx'
    basename = os.path.basename(filepath)
    filesize = os.path.getsize(filepath)

    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = sender
    msg['To'] = sender


    '''
    part_file = MIMEBase('application', 'octet-stream; name="{}"'.format(basename))
    part_file.set_payload(open(filepath, "rb").read())
    part_file.add_header('Content-Description', basename)
    part_file.add_header('Content-Disposition', 'attachment; filename="{}"; size={}'.format(basename, filesize))
    encoders.encode_base64(part_file)
    '''
    
    msg.attach(filepath)

    mail = smtplib.SMTP_SSL(server)
    mail.login(user, password)
    mail.sendmail(sender, recipients, msg.as_string())
    mail.quit()

            







send_email()






